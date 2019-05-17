from openpyxl import Workbook
import openpyxl
import os.path


class XLSX:
    """
    AUTHOR: WISROVI

    Esta es una clase para manipular archivos XLSX (la extensión xlsx es primordiar)
    en el constructor se pasan parametros de nombreArchivo, nombreHoja y en una tupla los nombres (string) de las variables

    Esta clase ofrece varias funciones claves:
    UsarOtraHoja(hojaUsar) #donde 'hojaUsar' es el nombre de la nueva hoja

    InsertarFila(valores) #donde 'valores' es una tupla del tamaño de la tupla del constructor


    """
    def __init__(self, nombreArchivo, nombreHoja=None, Columnas=()):
        self.nombreColumnas = Columnas
        self.nameFila = nombreArchivo
        if self.__VerificarExistenciaAchivo(nombreArchivo):
            self.book = openpyxl.load_workbook(nombreArchivo)
        else:
            self.book = Workbook()

        if nombreHoja == None:
            self.sheet = self.book.active
        else:
            if self.archivoExiste == False:
                self.CrearHoja(nombreHoja)
                self.BorrarHoja("Sheet")
            self.sheet = self.book.get_sheet_by_name(nombreHoja)

        if self.archivoExiste == False:
            self.InsertarFila(self.nombreColumnas)

        self.sheet.sheet_properties.tabColor = "0072BA"


    def __VerificarExistenciaAchivo(self, archivo):
        if os.path.isfile(archivo):
            self.archivoExiste=True
            return True
        else:
            self.archivoExiste=False
            return False




    def UsarOtraHoja(self, nombreHoja, Columnas):
        hojasActuales = self.VerHojasDocumento()
        if self.__existeHoja(nombreHoja):
            self.sheet = self.book.get_sheet_by_name(nombreHoja)
        else:
            self.CrearHoja(nombreHoja)
            self.sheet = self.book.get_sheet_by_name(nombreHoja)
            self.InsertarFila(Columnas)





    def __existeHoja(self, nombreHoja):
        hojasActuales = self.VerHojasDocumento()
        for hoja in hojasActuales:
            if hoja.__eq__(nombreHoja):
                return True
        return False



    def getCelda(self, celda):
        campo = self.sheet[str(celda)]
        return campo.value

    def getCelda(self, columna, fila):
        self.sheet.cell(fila,columna)









    def setCelda(self, celda, valor):
        self.sheet[str(celda)] = valor

    def setCelda(self, columna, fila, valor):
        self.sheet.cell(fila, columna, valor)








    def InsertarFila(self, VectorFila):
        self.sheet.append(VectorFila)


















    def CrearHoja(self, nuevaHoja):
        self.book.create_sheet(title=nuevaHoja)

    def BorrarHoja(self, nombreHoja):
        self.book.remove(self.book.get_sheet_by_name(nombreHoja))

    def VerHojasDocumento(self):
        return self.book.sheetnames

    def getTituloHoja(self):
        return self.sheet.title



    def LeerFilas(self):
        rows = self.sheet.rows
        contadorFilas = 0
        contadorColumnas = 0
        tamanoTupla = len(self.nombreColumnas)
        respuesta = "{"
        for row in rows:
            if contadorFilas > 0:
                respuesta += "'Fila" + str(contadorFilas) + "' : {"
            for cell in row:
                nombreCampo = ""
                if contadorColumnas < tamanoTupla:
                    nombreCampo = "'" + self.nombreColumnas[contadorColumnas] + "' : "
                else:
                    nombreCampo = "'NN' : "

                if contadorFilas > 0:
                    respuesta += nombreCampo + "'" + str(cell.value) + "'"
                    if contadorColumnas < self.sheet.max_column-1:
                        respuesta += ","
                contadorColumnas += 1

            if contadorFilas > 0:
                respuesta += "}"
                if contadorFilas < self.sheet.max_row-1:
                    respuesta += ","


            contadorColumnas = 0
            contadorFilas += 1

        respuesta += "}"
        return respuesta

    def ReadSizeSheet(self):
        return "{'FilaMin': '" + str(self.sheet.min_row) + "'," \
                "'FilaMax': '" + str(self.sheet.max_row) + "'," \
                "'ColumnaMin': '" + str(self.sheet.min_column) + "'," \
                "'ColumnaMax': '" + str(self.sheet.max_column) + "' }"




    def GuardarDatos(self):
        self.book.save(self.nameFila)
